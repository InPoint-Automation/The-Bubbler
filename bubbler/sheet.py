# Bubbler - Copyright (C) 2026 InPoint Automation Sp. z o.o.
# Licensed under the GNU General Public License v3 or later; see LICENSE.
#
# Writes inspection .xlsx from template, atomic save.

import base64
import os
import sys
from openpyxl import load_workbook

from .common import SHEET, FIRST_ROW, LAST_ROW

HEADER_FIELDS = [
    ("B3", "Part #"), ("E3", "Part Name / Nazwa"), ("H3", "FAIR #"),
    ("B4", "Drawing # / Rysunek"), ("E4", "Dwg Rev"), ("H4", "Part Rev"),
    ("B5", "PO #"), ("E5", "Material / Materiał"), ("H5", "Serial/Lot / Partia"),
    ("B6", "Inspector / Kontroler"), ("E6", "Date / Data"),
    ("H6", "FAI type"), ("J6", "Stage / Etap"),
]


def ensure_xlsx(path, company=""):
    if os.path.isfile(path):
        return False
    try:
        from .sheet_template import TEMPLATE_B64
    except ImportError:
        raise RuntimeError(
            "bubbler/sheet_template.py missing - run "
            "template/make_template_module.py or place the xlsx at %s" % path)
    with open(path, "wb") as f:
        f.write(base64.b64decode(TEMPLATE_B64))
    if company:
        try:
            wb = load_workbook(path)
            ws = wb[SHEET]
            ws["A1"] = "%s - Inspection Sheet / Karta kontroli" % company
            wb.save(path)
        except Exception as e:
            print("bubbler: could not set company header (%s)" % e,
                  file=sys.stderr)
    return True


class SheetWriter(object):
    COLS = (("A", "bubble"), ("B", "type"), ("C", "group"), ("D", "feature"),
            ("E", "nominal"), ("F", "tol_sym"), ("G", "tol_max"),
            ("H", "tol_min"), ("I", "pin"), ("J", "offset"),
            ("L", "measured"), ("N", "tier"), ("O", "gage"))
    NUMCOLS = {"E", "F", "G", "H", "I", "J"}

    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        if SHEET not in self.wb.sheetnames:
            raise ValueError("No '%s' sheet in %s" % (SHEET, path))
        self.ws = self.wb[SHEET]

    def get_header(self):
        return {cell: (self.ws[cell].value or "")
                for cell, _ in HEADER_FIELDS}

    def set_header(self, values):
        for cell, v in values.items():
            self.ws[cell] = v

    def next_row(self):
        for r in range(FIRST_ROW, LAST_ROW + 1):
            if self.ws["A%d" % r].value in (None, ""):
                return r
        return None

    def write_row(self, r, d):
        for col, key in self.COLS:
            v = d.get(key, None)
            if v in (None, ""):
                self.ws["%s%d" % (col, r)] = None
            elif col in self.NUMCOLS:
                try:
                    self.ws["%s%d" % (col, r)] = float(str(v).replace(",", "."))
                except ValueError:
                    self.ws["%s%d" % (col, r)] = str(v)
            elif col == "L":
                try:
                    self.ws["L%d" % r] = float(str(v).replace(",", "."))
                except ValueError:
                    self.ws["L%d" % r] = str(v)
            else:
                self.ws["%s%d" % (col, r)] = str(v)

    def clear_row(self, r):
        for col, _ in self.COLS:
            self.ws["%s%d" % (col, r)] = None

    def save(self):
        # write-then-replace, crash-safe
        tmp = self.path + ".tmp"
        self.wb.save(tmp)
        try:
            os.replace(tmp, self.path)
        except OSError:
            try:
                os.remove(tmp)
            except OSError:
                pass
            raise


